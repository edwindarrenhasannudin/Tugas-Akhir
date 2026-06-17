#!/usr/bin/env python3
"""
Export data lauk dan sayuran dari nutrition_pipeline.xlsx ke JSON untuk frontend React

Lokasi:
  - Input: ../nutrition/dataset/nutrition_pipeline.xlsx
  - Output: ../src/data/lauk_dataset.json, ../src/data/sayuran_dataset.json

Usage:
  python export_lauk_sayuran.py
"""

import openpyxl
import json
import sys
import os

def get_sheet_data(sheet):
    """Convert openpyxl sheet to list of dicts"""
    data = []
    headers = []
    
    # Get headers from first row
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Get data from remaining rows
    for row_idx in range(2, sheet.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            # Convert to float if numeric
            if isinstance(value, (int, float)) and header not in ['name', 'kategori', 'texture']:
                try:
                    row_data[header] = float(value) if value is not None else None
                except (ValueError, TypeError):
                    row_data[header] = value
            else:
                row_data[header] = value
        if any(row_data.values()):  # Only add non-empty rows
            data.append(row_data)
    
    return data

try:
    # Define paths relative to scripts folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, '../nutrition/dataset/nutrition_pipeline.xlsx')
    output_lauk = os.path.join(script_dir, '../src/data/lauk_dataset.json')
    output_sayuran = os.path.join(script_dir, '../src/data/sayuran_dataset.json')
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"❌ File tidak ditemukan: {input_file}")
        print(f"   Working directory: {os.getcwd()}")
        sys.exit(1)
    
    print("📂 Membaca data dari nutrition_pipeline.xlsx...")
    print(f"   Input: {input_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Get lauk and sayuran sheets
    lauk_sheet = wb['lauk']
    sayuran_sheet = wb['sayuran']
    
    # Convert to data
    lauk_data = get_sheet_data(lauk_sheet)
    sayuran_data = get_sheet_data(sayuran_sheet)
    
    print(f"✓ Lauk: {len(lauk_data)} rows")
    print(f"✓ Sayuran: {len(sayuran_data)} rows")
    
    # Ensure output directories exist
    os.makedirs(os.path.dirname(output_lauk), exist_ok=True)
    os.makedirs(os.path.dirname(output_sayuran), exist_ok=True)
    
    # Save to JSON files
    with open(output_lauk, 'w', encoding='utf-8') as f:
        json.dump(lauk_data, f, ensure_ascii=False, indent=2)
    print(f"✅ Saved: {output_lauk}")
    
    with open(output_sayuran, 'w', encoding='utf-8') as f:
        json.dump(sayuran_data, f, ensure_ascii=False, indent=2)
    print(f"✅ Saved: {output_sayuran}")
    
    # Print sample
    print(f"\n📊 Sample lauk (baris pertama):")
    if lauk_data:
        sample = lauk_data[0]
        print(f"   Name: {sample.get('name')}")
        print(f"   Kategori: {sample.get('kategori')}")
        print(f"   Texture: {sample.get('texture')}")
    
    wb.close()
    print("\n✨ Export complete!")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
