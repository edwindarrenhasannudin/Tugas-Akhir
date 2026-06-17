"""
scaling_nutrition.py - STEP 1 Pipeline
========================================
Membaca data nutrisi dari CSV, menyimpan dataset asli pada sheet terpisah,
dan menerapkan Min-Max Scaling pada kolom nutrisi.

Alur Pipeline (3 STEPS):
  STEP 1: scaling_nutrition.py (file ini)
    → Baca nutrition.csv
    → Min-Max Scaling pada kolom nutrisi
    → Output: sheet 'Nutrition Scaled'
  ↓
  STEP 2: onehot_encoding_nutrition.py
    → Klasifikasi texture & kategori
    → One-Hot Encoding
    → Output: sheet 'One-Hot Encoded'
  ↓
  STEP 3: pembagian.py (Merged pembersihan + pembagian)
    → Pembersihan data (hapus lainnya, keyword, strip, deduplikasi)
    → Pembagian per kategori (lauk, sayuran)
    → Output: sheet 'lauk' & 'sayuran'

File: nutrition_pipeline.xlsx
"""

import pandas as pd
import warnings
from sklearn.preprocessing import MinMaxScaler

warnings.filterwarnings('ignore')


# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def _format_excel_sheet(worksheet, header_color="4472C4"):
    """Format sheet Excel: header berwarna, border, auto-width, 4 desimal."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Format header
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Format data
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for idx, cell in enumerate(row, 1):
            cell.border = border
            # Kolom name: left-aligned
            if idx == 6:  # Assuming 'name' adalah kolom ke-6
                cell.alignment = left
            # Kolom nutrisi: center, 4 desimal
            elif idx in [2, 3, 4, 5]:
                cell.alignment = center
                if cell.value is not None:
                    try:
                        float(cell.value)
                        cell.number_format = '0.0000'
                    except (ValueError, TypeError):
                        pass

    # Auto-adjust column width
    for column in worksheet.columns:
        max_len = max((len(str(c.value or '')) for c in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)


def apply_minmax_scaling(df, columns_to_scale=None):
    """
    Menerapkan Min-Max Scaling pada kolom nutrisi.
    
    Min-Max Scaling: X_scaled = (X - X_min) / (X_max - X_min)
    Hasil akan bernilai antara 0-1
    
    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame yang akan di-scale
    columns_to_scale : list, optional
        Kolom yang akan di-scale. Default: ['calories', 'proteins', 'fat', 'carbohydrate']
    
    Returns:
    --------
    (df_scaled, scaler, scaling_info)
        - df_scaled: DataFrame dengan kolom yang sudah di-scale
        - scaler: MinMaxScaler object untuk referensi
        - scaling_info: Dictionary berisi info scaling tiap kolom
    """
    
    if columns_to_scale is None:
        columns_to_scale = ['calories', 'proteins', 'fat', 'carbohydrate']
    
    # Validasi kolom
    missing_cols = [col for col in columns_to_scale if col not in df.columns]
    if missing_cols:
        print(f"   WARNING: Kolom tidak ditemukan: {missing_cols}")
        columns_to_scale = [col for col in columns_to_scale if col in df.columns]
    
    if not columns_to_scale:
        print("   ERROR: Tidak ada kolom yang valid untuk di-scale")
        return df, None, {}
    
    df_scaled = df.copy()
    
    # Terapkan MinMaxScaler
    scaler = MinMaxScaler()
    scaled_values = scaler.fit_transform(df[columns_to_scale])
    df_scaled[columns_to_scale] = scaled_values
    
    # Hitung info scaling
    scaling_info = {}
    for idx, col in enumerate(columns_to_scale):
        scaling_info[col] = {
            'min': df[col].min(),
            'max': df[col].max(),
            'range': df[col].max() - df[col].min(),
        }
    
    return df_scaled, scaler, scaling_info


def _print_scaling_report(df_original, df_scaled, scaling_info, columns_to_scale):
    """Cetak laporan detail tentang proses scaling ke terminal."""
    
    print("\n" + "-" * 80)
    print("[REPORT] HASIL MIN-MAX SCALING:")
    print("-" * 80)
    
    print("\nRumus: X_scaled = (X - X_min) / (X_max - X_min)")
    print("Hasil: Nilai antara 0 (minimum) hingga 1 (maksimum)\n")
    
    for col in columns_to_scale:
        if col in scaling_info:
            info = scaling_info[col]
            print(f"[{col.upper()}]")
            print(f"   Nilai minimum: {info['min']:.4f}")
            print(f"   Nilai maksimum: {info['max']:.4f}")
            print(f"   Range: {info['range']:.4f}")
            print(f"   Scaled min: {df_scaled[col].min():.4f}")
            print(f"   Scaled max: {df_scaled[col].max():.4f}")
            print()
    
    # Preview data
    print("-" * 80)
    print("[PREVIEW] PERBANDINGAN DATA SEBELUM DAN SESUDAH SCALING (5 Baris Pertama):")
    print("-" * 80)
    print("\nSEBELUM SCALING:")
    print(df_original[['name'] + columns_to_scale].head(5).to_string(index=False))
    print("\nSESUDAH SCALING:")
    print(df_scaled[['name'] + columns_to_scale].head(5).to_string(index=False))


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import os
    
    print("=" * 80)
    print("[INFO] MIN-MAX SCALING - STEP 1 (Pipeline)")
    print("=" * 80)
    
    # Path
    input_csv = "dataset/nutrition.csv"
    output_file = "dataset/nutrition_pipeline.xlsx"
    
    # Cek file input
    if not os.path.exists(input_csv):
        print(f"\n[ERROR] File '{input_csv}' tidak ditemukan!")
        print("   Pastikan file nutrition.csv ada di folder nutrition/dataset")
        exit(1)
    
    try:
        # ---- Step 1a: Baca data dari CSV ----
        print(f"\n[READ] Membaca file: {input_csv}")
        df = pd.read_csv(input_csv)
        print(f"[OK] Total data: {len(df)} bahan makanan")
        print(f"[OK] Kolom yang tersedia: {list(df.columns)}\n")
        
        df_original = df.copy()
        
        # ---- Step 1b: Simpan Dataset Asli ----
        print("[WRITE] Menyimpan ke sheet 'Dataset Asli'...")
        
        # Cek apakah file sudah ada dan valid (bukan file kosong/corrupt)
        file_is_valid = False
        if os.path.exists(output_file):
            if os.path.getsize(output_file) > 0:
                try:
                    from openpyxl import load_workbook
                    load_workbook(output_file).close()
                    file_is_valid = True
                except Exception:
                    file_is_valid = False
            
            if not file_is_valid:
                print(f"   WARNING: File '{output_file}' rusak/kosong, akan dibuat ulang.")
                os.remove(output_file)
        
        # Gunakan mode append jika file sudah ada dan valid
        if file_is_valid:
            print(f"   INFO: Updating sheet dalam {output_file}")
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Dataset Asli', index=False)
                _format_excel_sheet(writer.sheets['Dataset Asli'])
        else:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            print(f"   INFO: Membuat file baru: {output_file}")
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Dataset Asli', index=False)
                _format_excel_sheet(writer.sheets['Dataset Asli'])
        
        print("[OK] Sheet 'Dataset Asli' berhasil disimpan!\n")
        
        # ---- Step 1c: Terapkan Min-Max Scaling ----
        print("=" * 80)
        print("[STEP 1c] PENERAPAN MIN-MAX SCALING")
        print("=" * 80)
        print("   Metode: Min-Max Normalization")
        print("   Target kolom: calories, proteins, fat, carbohydrate")
        print("-" * 80)
        
        columns_to_scale = ['calories', 'proteins', 'fat', 'carbohydrate']
        df_scaled, scaler, scaling_info = apply_minmax_scaling(df, columns_to_scale)
        
        _print_scaling_report(df_original, df_scaled, scaling_info, columns_to_scale)
        
        # ---- Step 1d: Simpan hasil scaling ----
        print(f"\n[WRITE] Menambahkan sheet 'Nutrition Scaled'...")
        
        df_scaled = df_scaled.drop(columns=['image'], errors='ignore')
        
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_scaled.to_excel(writer, sheet_name='Nutrition Scaled', index=False)
            _format_excel_sheet(writer.sheets['Nutrition Scaled'])
        
        print("[OK] Sheet 'Nutrition Scaled' berhasil ditambahkan!\n")
        
        # ---- Ringkasan ----
        print("=" * 80)
        print("[RESULT] RINGKASAN MIN-MAX SCALING")
        print("=" * 80)
        print(f"\n[OK] Data asli: {len(df)} bahan makanan")
        print(f"[OK] Data terskalakan: {len(df_scaled)} bahan makanan")
        print(f"[OK] Kolom asli: {df.shape[1]}")
        print(f"[OK] Kolom setelah scaling: {df_scaled.shape[1]} (sama, hanya nilai yang berubah)")
        
        print(f"\n[KOLOM TERSKALAKAN]:")
        for col in columns_to_scale:
            print(f"   - {col}")
        
        print(f"\n[KOLOM TIDAK TERSKALAKAN]:")
        non_scaled = [col for col in df.columns if col not in columns_to_scale]
        for col in non_scaled:
            print(f"   - {col}")
        
        print("\n[DATA] Output tersimpan di:")
        print(f"   [Sheet] 'Dataset Asli': Data original dari CSV")
        print(f"   [Sheet] 'Nutrition Scaled': Data setelah Min-Max Scaling")
        print(f"   [File] {output_file}")
        
        print("\n[DONE] STEP 1 SELESAI: Min-Max Scaling berhasil diterapkan!")
        print("\n" + "=" * 80)
        print("NEXT STEP: Jalankan onehot_encoding_nutrition.py")
        print("=" * 80)
        
    except FileNotFoundError as e:
        print(f"[ERROR] File tidak ditemukan: {e}")
    except KeyError as e:
        print(f"[ERROR] Kolom tidak ditemukan: {e}")
    except Exception as e:
        print(f"[ERROR] Error: {e}")
        import traceback
        traceback.print_exc()