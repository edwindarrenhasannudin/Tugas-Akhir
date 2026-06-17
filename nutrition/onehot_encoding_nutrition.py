"""
onehot_encoding_nutrition.py - STEP 2 Pipeline
================================================
Mengklasifikasikan bahan makanan berdasarkan nama, lalu
menerapkan One-Hot Encoding pada kolom texture & kategori.

Alur Pipeline (3 STEPS):
  STEP 1: scaling_nutrition.py
    → Baca nutrition.csv
    → Min-Max Scaling pada kolom nutrisi
    → Output: sheet 'Nutrition Scaled'
  ↓
  STEP 2: onehot_encoding_nutrition.py (file ini)
    → Klasifikasi texture & kategori
    → One-Hot Encoding
    → Output: sheet 'One-Hot Encoded'
  ↓
  STEP 3: pembagian.py (Merged pembersihan + pembagian)
    → Pembersihan data (hapus lainnya, keyword, strip, deduplikasi)
    → Pembagian per kategori (lauk, sayuran)
    → Output: sheet 'lauk' & 'sayuran'

File: nutrition_pipeline.xlsx
"""

import re
import pandas as pd
from collections import defaultdict

import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# KONFIGURASI KEYWORD
# ============================================================================

# Exact mapping untuk kasus-kasus khusus yang sudah diketahui
EXACT_MAPPINGS = {
    'bayam':              ('lembut', 'sayuran'),
    'keripik bayam':      ('renyah', 'sayuran'),
    'susu kental manis':  ('kental', 'susu'),
    'susu segar':         ('cair',   'susu'),
    'minyak kelapa':      ('kental', 'minyak/lemak'),
    'air kelapa':         ('cair',   'buah'),
    'santan kelapa':      ('cair',   'buah'),
    'telur ayam':         ('cair',   'lauk'),
    'telur bebek':        ('cair',   'lauk'),
    'daun bayam':         ('lembut', 'sayuran'),
    'labu air':           ('padat',  'sayuran'),
}

# Urutan prioritas (untuk tie-breaking saat skor sama)
TEXTURE_PRIORITY = ['cair', 'kental', 'lembut', 'padat', 'renyah', 'bubuk', 'netral']
CATEGORY_PRIORITY = [
    'telur', 'susu', 'minyak/lemak', 'konfeksioneri',
    'bumbu_dan_rempah', 'lauk', 'kacang_kacangan',
    'sayuran', 'buah', 'serelia', 'umbi_berpati', 'lainnya',
]

# Keyword -> texture
TEXTURE_KEYWORDS = {
    'padat': [
        'daging', 'ayam', 'sapi', 'kambing', 'dendeng', 'anggur', 'arrowroot',
        'kandis', 'asam payak', 'bagea', 'beef', 'biji', 'brem', 'baligo', 'batang',
        'rotan', 'rukam', 'bungkil', 'cakalang', 'cantel', 'cue', 'hangop', 'jahe',
        'jampang', 'jawawut', 'lobak', 'karawila', 'kemiri', 'kenari', 'ketapang',
        'kluwek', 'geplak', 'lapis', 'sente', 'uwi', 'beras', 'bit', 'oncom', 'gadung',
        'jagung', 'jantung', 'kaburan', 'kentang', 'keribang', 'rebung', 'kunyit',
        'lepok', 'terasi', 'ubi', 'wortel', 'babat', 'batatas', 'bebek', 'komba',
        'boros', 'buah merah', 'ruruhi', 'gatot', 'permen', 'geblek', 'gembili', 'gula',
        'kapurung', 'kelapa hutan', 'ketela', 'ketumbar', 'kranji', 'kundur', 'labu',
        'cengkeh', 'kabuto', 'kawista', 'kerbau', 'lokan', 'melinjo', 'takwa',
        'talas', 'udang', 'bekasang', 'kakap', 'mujair', 'tempe', 'waluh', 
    ],
    'lembut': [
        'haruwan', 'bubur', 'kukus', 'rebus', 'nasi', 'alpukat', 'arbei', 'ares',
        'arwan', 'es', 'bawang', 'betok', 'botok', 'bakpia', 'bakung', 'bantal',
        'barongko', 'atung', 'negri', 'nona', 'cammetutu', 'cempedak', 'fillet',
        'kelewih', 'kesemek', 'kwaci', 'selat', 'rajungan', 'sawo', 'tempoya',
        'tempuyak', 'mentega', 'ampas', 'daun', 'daun ubi', 'getuk', 'gudeg', 'mie',
        'kangkung', 'keju', 'ketoprak', 'krim', 'kepiting', 'mi', 'brongkos', 'nanas',
        'roti', 'sagu', 'sukun', 'bayam', 'tahu', 'batar', 'bihun', 'bika', 'naga',
        'tuppa', 'kelor', 'bunga', 'buntil', 'tape', 'carica', 'duku', 'durian',
        'duwet', 'embacang', 'gambas', 'gurandil', 'jali', 'jamur', 'kambose',
        'kaparende', 'kapusa', 'kelepon', 'kemang', 'kerang', 'ketan', 'kokosan',
        'kulit melinjo', 'kwark', 'turi', 'lamtoro', 'lantar', 'lema', 'lontar',
        'makaroni', 'mangga', 'manggis', 'margarin', 'masekat', 'matoa', 'melon',
        'menteng', 'misoa', 'nangka', 'oyek', 'papeda', 'pepaya', 'pisang', 'terong',
        'terung', 'pulut', 'putu', 'rambutan', 'sardines', 'tapai', 'serimping',
        'sirsak', 'spaghetti', 'srikaya', 'suweg', 'tiwul', 'tomat', 'yangko',
        'ikan', 'cumi',
    ],
    'cair': [
        'sop', 'santan', 'sayur', 'cuka', 'air', 'gulai', 'jeruk', 'lemon', 'lemonade',
        'telur', 'leunca buah', 'rujak', 'semangka', 'susu', 'seduh', 'teh', 'tebu',
    ],
    'kental': [
        'kental', 'kecap', 'madu', 'rusip', 'taoco', 'yoghurt', 'sirup', 'selai',
        'melase', 'petis', 'saos', 'tauco', 'minyak', 'tauji',
    ],
    'renyah': [
        'akar tonjong', 'anyang', 'kerupuk', 'keripik', 'rempeyek', 'apel', 'emping',
        'biskuit', 'aletoge', 'andewi', 'bengkuang', 'biwah', 'bakwan', 'beberuk',
        'belimbing', 'kom', 'cabai', 'kelenting', 'eceng', 'enting', 'gatep', 'jengkol',
        'jotang', 'kabau', 'komak', 'kacang', 'koro', 'kalakai', 'kadada', 'kapri',
        'kecombrang', 'keremes', 'kerokot', 'gemblong', 'genjer', 'krokot', 'nopia',
        'paria', 'pete', 'peterseli', 'rimbang', 'tekokak', 'toge', 'wijen', 'widaran',
        'buncis', 'caisin', 'encung', 'gandaria', 'paku', 'ganyong', 'jambu',
        'jambu air', 'selada air', 'selada', 'karedok', 'kecipir', 'ketimun', 'markisa',
        'mostarda', 'pastel', 'salak', 'sawi', 'seledri', 'taoge', 
        'teri balado', 'kedondong', 'bunga pepaya', 'cap cai', 'erbis', 'japilus',
        'kool', 'kucai', 'laksa', 'rebon',
    ],
    'bubuk': [
        'tepung', 'bubuk', 'serbuk', 'kopi', 'coklat bubuk', 'koya', 'maizena', 'merica',
    ],
}

# Keyword -> kategori
CATEGORY_KEYWORDS = {
    'serelia': [
        'beras', 'cantel', 'jali', 'jawawut', 'jampang', 'nasi', 'tapai',
        'tepung', 'bihun', 'ketupat', 'ketan', 'maizena', 'makaroni', 'misoa', 'roti',
        'apem', 'biskuit', 'bakpia', 'bakwan', 'bantal', 'batar daan', 'bika', 'brem',
        'bubur', 'cake tape', 'dodol', 'gemblong', 'gendar', 'intip', 'japilus',
        'kambose', 'kapusa', 'kelepon', 'kue', 'ketoprak', 'koya', 'laksa',
        'lapis legit', 'putu mayang', 'lupis', 'martabak', 'masekat', 'nopia',
        'onde-onde', 'pastel', 'pulut', 'pundut', 'putri selat', 'renggi', 'sarimuka',
        'spaghetti', 'suwir-suwir', 'tipa-tipa', 'wajit', 'widaran', 'wingko', 'yangko',
    ],
    'umbi_berpati': [
        'arrowroot', 'batatas', 'belitung', 'bengkuang', 'bentul', 'gadeng/gadung',
        'gadung', 'ganyong', 'gembili', 'hofa/ubi', 'kentang', 'keribang', 'ketela',
        'lepok/ubi', 'sagu', 'sente', 'talas', 'umbi uwi', 'uwi', 'ubi', 'suweg',
        'bagea', 'biji', 'cassavastick', 'ceriping', 'gatot', 'getuk', 'geblek',
        'gurandil', 'kabuto', 'kapurung', 'kecimpring', 'keremes', 'keripik', 'kerupuk',
        'oyek', 'papeda', 'rasbi', 'rasi', 'serimping', 'tiwul',
    ],
    'kacang_kacangan': [
        'kacang', 'kenari', 'komak', 'koro', 'lamtoro', 'wijen', 'ampas', 'pepea',
        'bungkil', 'emping', 'enting-enting', 'geplak', 'kembang', 'kwaci', 'oncom',
        'melinjo', 'lebui', 'kedelai', 'takwa', 'tauco', 'taoco', 'tauji',
    ],
    'sayuran': [
        'kangkung', 'wortel', 'brokoli', 'kubis', 'sawi', 'selada', 'daun', 'sayur',
        'kentang', 'tomat', 'akar tonjong', 'aletoge', 'andaliman', 'andewi', 'bakung',
        'buah kelor', 'kool kembang', 'buncis', 'caisin', 'ketimun', 'labu', 'rebung',
        'lobak', 'selada', 'terong', 'gambas', 'bawang', 'kucai', 'jagung', 'jamur', 
        'jengkol',
    ],
    'buah': [
        'apel', 'jeruk', 'pisang', 'mangga', 'anggur', 'pepaya', 'semangka', 'durian',
        'nanas', 'duku', 'alpukat', 'belimbing', 'buah', 'kelapa',
    ],
    'lauk': [
        'daging', 'ayam', 'ikan', 'sapi', 'kambing', 'bebek', 'rusa', 'kerbau', 'udang',
        'cumi', 'kepiting', 'kerang', 'telur', 'tahu', 'tempe',
    ],
    'susu':            ['susu', 'keju', 'yoghurt', 'mentega', 'krim'],
    'minyak/lemak':    ['minyak', 'lemak'],
    'bumbu_dan_rempah': [
        'garam', 'gula', 'merica', 'lada', 'kunyit', 'jahe', 'cengkeh',
        'pala', 'saos',
    ],
    'konfeksioneri':   ['gula', 'permen', 'coklat'],
    'lainnya':         [],
}


# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def _word_match(keyword, text):
    """
    Match keyword sebagai kata utuh menggunakan regex word boundary.
    Mencegah false positive seperti 'ayam' cocok pada 'bayam'.
    """
    pattern = r'\b' + re.escape(keyword.strip()) + r'\b'
    return bool(re.search(pattern, text))


def _compute_scores(name_lower, keywords_dict, priority_list):
    """
    Hitung skor matching dengan weighted scoring.
    Keyword multi-kata mendapat skor lebih tinggi (bobot x2).

    Returns: (best_match, matched_keywords_dict) atau (None, {}) jika tidak ada match.
    """
    scores = defaultdict(float)
    matched = defaultdict(list)

    for category, keywords in keywords_dict.items():
        for kw in sorted(keywords, key=len, reverse=True):
            kw = kw.strip()
            if kw and _word_match(kw, name_lower):
                word_count = len(kw.split())
                scores[category] += (word_count * 2) if word_count > 1 else 1
                matched[category].append(kw)

    if not scores:
        return None, {}

    best = sorted(
        scores.items(),
        key=lambda x: (-x[1], priority_list.index(x[0]) if x[0] in priority_list else 999)
    )[0][0]

    return best, matched


def _format_excel_sheet(worksheet):
    """Format sheet Excel: header biru, border, auto-width, 4 desimal."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Header
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Data
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center
            if cell.column > 1:
                try:
                    if cell.value is not None and float(cell.value):
                        cell.number_format = '0.0000'
                except (ValueError, TypeError):
                    pass

    # Auto-width
    for column in worksheet.columns:
        max_len = max((len(str(c.value or '')) for c in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)


# ============================================================================
# FUNGSI UTAMA
# ============================================================================

def apply_onehot_encoding(df, categorical_columns=None):
    """
    Menerapkan One-Hot Encoding pada kolom kategori.

    Returns: (df_encoded, encoding_mapping)
    """
    if categorical_columns is None:
        categorical_columns = ['texture', 'kategori']

    df_encoded = df.copy()
    encoding_mapping = {}

    for col in categorical_columns:
        if col not in df_encoded.columns:
            print(f"[WARNING] Kolom '{col}' tidak ditemukan dalam DataFrame")
            continue

        unique_values = df_encoded[col].unique()
        print(f"\n[COLUMN] Kolom '{col}':")
        print(f"   - Total kategori: {len(unique_values)}")
        print(f"   - Kategori: {sorted(unique_values)}")

        one_hot = pd.get_dummies(df_encoded[col], prefix=col, dtype='int')

        encoding_mapping[col] = {
            'categories': sorted(unique_values),
            'encoded_columns': one_hot.columns.tolist(),
        }

        df_encoded = pd.concat([df_encoded, one_hot], axis=1)

    return df_encoded, encoding_mapping


def classify_food_item(name):
    """
    Klasifikasi makanan -> (texture, kategori)

    Pendekatan berlapis:
      Layer 1: Exact mapping (kasus khusus)
      Layer 2: Word-boundary matching + weighted scoring
      Layer 3: Override rules (konflik umum)
    """
    name_lower = re.sub(r'[-_/]', ' ', name.lower()).strip()

    # --- Layer 1: Exact mapping ---
    if name_lower in EXACT_MAPPINGS:
        return EXACT_MAPPINGS[name_lower]

    # --- Layer 2: Word-boundary matching + weighted scoring ---
    best_texture, _ = _compute_scores(name_lower, TEXTURE_KEYWORDS, TEXTURE_PRIORITY)
    best_category, _ = _compute_scores(name_lower, CATEGORY_KEYWORDS, CATEGORY_PRIORITY)

    if best_texture is None:
        best_texture = 'netral'
    if best_category is None:
        best_category = 'lainnya'

    # --- Layer 3: Override rules ---
    if _word_match('susu', name_lower):
        best_texture = 'kental' if _word_match('kental', name_lower) else 'cair'
        best_category = 'susu'

    if _word_match('kuah', name_lower) or _word_match('sop', name_lower) or _word_match('jus', name_lower):
        best_texture = 'cair'

    if _word_match('goreng', name_lower) or _word_match('keripik', name_lower):
        best_texture = 'renyah'

    if _word_match('bubuk', name_lower) or _word_match('tepung', name_lower):
        best_texture = 'bubuk'

    if _word_match('kacang', name_lower) and _word_match('telur', name_lower):
        best_texture = 'renyah'
        best_category = 'kacang_kacangan'

    if _word_match('bayam', name_lower):
        best_category = 'sayuran'
        if not _word_match('keripik', name_lower):
            best_texture = 'lembut'

    if _word_match('kelapa', name_lower):
        best_category = 'buah'
        if not _word_match('minyak', name_lower):
            best_texture = 'padat'

    return best_texture, best_category


# ============================================================================
# FUNGSI TAMPILAN TERMINAL
# ============================================================================

def _print_classification_report(df, methods, total):
    """Cetak preview, distribusi texture, dan distribusi kategori ke terminal."""

    # Preview 15 sampel acak
    print("\n" + "-" * 80)
    print("[PREVIEW] HASIL KLASIFIKASI (15 sampel acak):")
    print("-" * 80)
    print(f"   {'No':<4} {'Nama Bahan':<35} {'Texture':<12} {'Kategori':<20} {'Metode'}")
    print(f"   {'-'*4} {'-'*35} {'-'*12} {'-'*20} {'-'*25}")

    sample_df = df.sample(n=min(15, len(df)), random_state=42)
    for idx, (i, row) in enumerate(sample_df.iterrows(), 1):
        m = methods[i] if i < len(methods) else "Word Boundary"
        print(f"   {idx:<4} {str(row['name']):<35} {row['texture']:<12} {row['kategori']:<20} {m}")

    # Distribusi texture
    print("\n" + "-" * 80)
    print("[DIST] DISTRIBUSI TEXTURE:")
    print("-" * 80)
    for label, count in df['texture'].value_counts().items():
        bar = "#" * (count // 20) + "." * max(0, 30 - count // 20)
        print(f"   {label:<10} {bar} {count:>5} ({count/total*100:>5.1f}%)")

    # Distribusi kategori
    print("\n" + "-" * 80)
    print("[DIST] DISTRIBUSI KATEGORI:")
    print("-" * 80)
    for label, count in df['kategori'].value_counts().items():
        bar = "#" * (count // 20) + "." * max(0, 30 - count // 20)
        print(f"   {label:<20} {bar} {count:>5} ({count/total*100:>5.1f}%)")


def _classify_all(df):
    """Klasifikasi semua baris, tampilkan progress bar, return methods list."""
    total = len(df)
    results = []
    methods = []

    for _, row in df.iterrows():
        name = row['name']
        normalized = re.sub(r'[-_/]', ' ', name.lower()).strip()
        method = "Exact Mapping" if normalized in EXACT_MAPPINGS else "Word Boundary + Scoring"

        texture, kategori = classify_food_item(name)
        results.append((texture, kategori))
        methods.append(method)

        # Progress setiap 200 item
        count = len(results)
        if count % 200 == 0 or count == total:
            pct = count / total * 100
            filled = int(pct // 5)
            bar = "#" * filled + "." * (20 - filled)
            print(f"\r   [{bar}] {count}/{total} ({pct:.0f}%) item diklasifikasi", end="", flush=True)

    print()  # newline setelah progress bar

    df['texture'] = [r[0] for r in results]
    df['kategori'] = [r[1] for r in results]

    return methods


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import os

    print("=" * 80)
    print("[INFO] ONE-HOT ENCODING - STEP 2 (Pipeline)")
    print("=" * 80)

    input_file = "dataset/nutrition_pipeline.xlsx"

    if not os.path.exists(input_file):
        print(f"[ERROR] File '{input_file}' tidak ditemukan!")
        print("   Pastikan Anda telah menjalankan scaling_nutrition.py terlebih dahulu")
        exit(1)

    try:
        # ---- Baca data ----
        print(f"\n[READ] Membaca file: {input_file}")
        df = pd.read_excel(input_file, sheet_name='Nutrition Scaled')
        print(f"[OK] Total data: {len(df)} bahan makanan")
        print(f"[OK] Kolom yang tersedia: {list(df.columns)}\n")

        # ---- Step 2a: Klasifikasi ----
        if 'texture' not in df.columns or 'kategori' not in df.columns:
            print("[WARNING] Kolom texture/kategori tidak ditemukan. Menambahkannya...")
            if 'name' not in df.columns:
                print("[ERROR] Kolom 'name' tidak ditemukan!")
                exit(1)

            print("\n" + "=" * 80)
            print("[STEP 2a] KLASIFIKASI TEXTURE & KATEGORI")
            print("=" * 80)
            print("   Metode: Word Boundary Matching + Weighted Scoring")
            print("   Layer 1: Exact Mapping (kasus khusus yang sudah diketahui)")
            print("   Layer 2: Keyword Matching dengan regex word boundary (\\b)")
            print("   Layer 3: Override Rules (konflik umum)")
            print("-" * 80)

            methods = _classify_all(df)
            _print_classification_report(df, methods, len(df))

            print("\n" + "-" * 80)
            print(f"[OK] Klasifikasi selesai: {len(df)} bahan makanan berhasil diklasifikasi!")
            print("-" * 80)

        # ---- Step 2b: One-Hot Encoding ----
        print("\n" + "=" * 80)
        print("[STEP 2b] ONE-HOT ENCODING")
        print("=" * 80)
        print("   Mengubah kolom kategorikal menjadi representasi binary (0/1)")
        print("-" * 80)

        df_encoded, encoding_mapping = apply_onehot_encoding(df)
        print("[OK] One-Hot Encoding selesai!")

        # ---- Simpan ke Excel ----
        print(f"\n[WRITE] Menambahkan sheet baru ke {input_file}...")

        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_encoded.to_excel(writer, sheet_name='One-Hot Encoded', index=False)
            _format_excel_sheet(writer.sheets['One-Hot Encoded'])

        print("[OK] Sheet 'One-Hot Encoded' berhasil ditambahkan!\n")

        # ---- Ringkasan ----
        print("=" * 80)
        print("[RESULT] RINGKASAN ONE-HOT ENCODING")
        print("=" * 80)
        print(f"\n[OK] Data: {len(df_encoded)} bahan makanan")
        print(f"[OK] Kolom asli: {len(df.columns)}")
        print(f"[OK] Kolom setelah one-hot: {len(df_encoded.columns)}")
        print(f"[OK] Kolom baru ditambahkan: {len(df_encoded.columns) - len(df.columns)}")

        print(f"\n[ENCODING] One-Hot Encoding diterapkan pada:")
        for col, info in encoding_mapping.items():
            print(f"   - {col}: {len(info['categories'])} kategori")

        print("\n[DONE] STEP 2 SELESAI: One-Hot Encoding berhasil disimpan!")
        print(f"[FILE] File: {input_file}")
        print("\n" + "=" * 80)
        print("NEXT STEP: Jalankan pembagian.py")
        print("=" * 80)

    except FileNotFoundError as e:
        print(f"[ERROR] File tidak ditemukan: {e}")
    except KeyError as e:
        print(f"[ERROR] Kolom tidak ditemukan: {e}")
    except Exception as e:
        print(f"[ERROR] Error: {e}")
        import traceback
        traceback.print_exc()