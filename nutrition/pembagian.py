"""
pembagian.py - STEP 3-4 Pipeline (MERGED)
=========================================
Menggabungkan pembersihan.py (STEP 3) dan pembagian.py (STEP 4).
Membersihkan + membagi data berdasarkan kategori dalam satu tahap.

Fitur:
  1. Membersihkan data hasil One-Hot Encoding
  2. Membagi berdasarkan kategori (lauk, sayuran)
  3. Output langsung ke sheet lauk & sayuran
  4. Tidak ada sheet intermediate 'Data Cleaned'

Alur Pipeline:
  scaling_nutrition.py (Step 1) 
  -> onehot_encoding_nutrition.py (Step 2) 
  -> pembagian.py (Step 3-4) ← file ini (MERGED)
"""

import re
import os
import pandas as pd
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# KONFIGURASI PEMBERSIHAN (dari pembersihan.py)
# ============================================================================

# Kata-kata yang akan DIHILANGKAN dari nama bahan di kolom 'name'.
STRIP_WORDS = [
    'segar', 'daging', 'pasar', 'tambak',
]

# Kata kunci untuk menghapus bahan makanan (case-insensitive)
KEYWORDS_TO_REMOVE = [
    'andaliman', 'pohon', 'babi', 'anjing', 'kuda', 'burung', 'angsa', 
    'belut', 'katak', 'keong', 'kodok', 'kura-kura', 'ulat sagu', 'ham', 'hiu', 
    'kotiu', 'buaya', 'dodol', 'penyu', 'masakan', 'goreng', 'kukus', 'lodeh',
    'ginjal', 'sosis', 'lilin', 'pempek', 'kue', 'martabak', 'otak', 'ular', 
    'purundawa', 'putri malu', 'purut', 'sate', 'sarimuka', 'tekwan', 'tinira', 
    'camilan', 'ketoprak', 'rusa', 'soto', 'anak', 'dideh/darah', 'hati', 'alabio', 
    'belibis', 'asap', 'dendeng', 'dideh', 'kering', 'bakar', 'jantung', 'perut',
    'mentah', 'hitam', 'kacangan', 'kawalinya', 'kima', 'lidah', 'nasu', 
    'betok', 'telan', 'bubuk', 'tepung', 'kerupuk', 'abon', 'gemuk', 'kornet',
    'kurus', 'lemak', 'kambing daging', 'kerbau daging', 'akar', 'asinan', 
    'rebus', 'sop', 'umbut', 'keripik', 'tondano', 'bader', 'balong', 'bambangan', 
    'baung', 'bekasang', 'belida', 'beunteur', 'biawan', 'bili', 'bubara', 'bulan-bulan', 
    'kakatua', 'kapar', 'katombo', 'layur', 'lehoma', 'lemuru', 'malalugis', 'pepes', 
    'mayong', 'oci', 'pepetek', 'sale', 'saluang', 'selar', 'sepat', 'sidat', 'sunu',
    'tahuman', 'tarmon', 'tembang', 'tempahas', 'terbang', 'titang', 'turi', 'peda', 'petis',
    'pisang', 'pindang', 'daleman', 'keleponan', 'sapi usus', 'tumis', 'sayur', 'belimbing',
    'ketupat', 'koro', 'lamtoro', 'tahu telur',  'ceplok', 'dadar', 'kampung', 'usus', 'liver',
    'ceplok', 'asin', 'terubuk', 'bongkrek', 'gembus', 'kacang', 'makanan', 'asam', 'madura',
    'hintalo', 'ampas', 'mie', 'kelinci', 'rebon', 'babat', 'galah', 'bagian', 'besar',
    'gulai', 'bandeng', 'banjar', 'baronang', 'cakalang', 'ikan daun', 'bawal', 'gabus', 'kakap', 
    'kembung', 'layang', 'mas', 'mujair', 'patin', 'sarden', 'teri', 'tongkol', 'kembang tahu', 
    'moon', 'ekor', 'labu kuning', 'kedelai', 'buntil', 'bangun-bangun', 'daun bawang merah', 
    'bakung', 'bebuas', 'belem', 'beluntas', 'cincau', 'gandaria', 'gelang', 'gunda', 'muda', 
    'jampang', 'jawaw', 'jonghe', 'karet', 'kenikir', 'kesum', 'daun kol sawi', 'kubis', 'kumak',
    'daun labu', 'leilem', 'leunca', 'lobak', 'lompong', 'mangkokan', 'ambon', 'ndusuk', 'daun oyong', 
    'pakis', 'paku', 'pangisegar', 'cina', 'poh-pohan', 'selasih', 'semanggi', 'simpur', 'singkil', 
    'ampenan', 'kopang', 'sintrong', 'tespong', 'talas', 'tales', 'ubi', 'grontol', 'giling', 'pipil', 
    'putih', 'titi', 'encik', 'sagu', 'katul', 'mostarda', 'nasi', 'pelecing', 'prey', 'taiwan', 'tanah',
    'semur', 'belanda', 'juice', 'andewi', 'waluh',
]

# ============================================================================
# KONFIGURASI PEMBAGIAN (dari pembagian.py)
# ============================================================================

# Kategori yang ingin diambil dan dijadikan sheet terpisah
TARGET_CATEGORIES = ['lauk', 'sayuran']


# ============================================================================
# FUNGSI UTILITAS
# ============================================================================

def strip_words_from_name(name):
    """
    Bersihkan nama bahan dengan menghilangkan kata-kata di STRIP_WORDS.
    
    Returns: (str_cleaned, list_of_removed_words)
    """
    if pd.isna(name):
        return name, []
    
    cleaned = str(name)
    removed = []
    
    for word in STRIP_WORDS:
        pattern = re.compile(r'\b' + re.escape(word) + r'\b', re.IGNORECASE)
        if pattern.search(cleaned):
            cleaned = pattern.sub('', cleaned)
            removed.append(word)
    
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned, removed


def format_excel_sheet(worksheet, header_color="27AE60"):
    """Format Excel sheet dengan styling profesional."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            try:
                if cell.value is not None:
                    float(cell.value)
                    cell.alignment = center
                    cell.number_format = '0.0000'
            except (ValueError, TypeError):
                cell.alignment = left

    for column in worksheet.columns:
        max_length = max((len(str(c.value or '')) for c in column), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


# ============================================================
# MAIN PROCESS
# ============================================================

print("=" * 80)
print("[STEP 3] PEMBERSIHAN + PEMBAGIAN DATA (Pipeline Merged)")
print("=" * 80)

# Baca file Excel dari STEP 2
input_file = 'dataset/nutrition_pipeline.xlsx'
print(f"\n[INFO] Membaca file: {input_file}")

if not os.path.exists(input_file):
    print(f"❌ File '{input_file}' tidak ditemukan!")
    print("   Pastikan Anda telah menjalankan onehot_encoding_nutrition.py terlebih dahulu")
    exit(1)

try:
    df = pd.read_excel(input_file, sheet_name='One-Hot Encoded')
    print(f"   [OK] Membaca sheet 'One-Hot Encoded'")
except Exception:
    print("   ⚠️ Sheet 'One-Hot Encoded' tidak ditemukan. Mencari sheet alternatif...")
    try:
        df = pd.read_excel(input_file, sheet_name='Nutrition Scaled')
        print(f"   [OK] Membaca sheet 'Nutrition Scaled' (fallback)")
    except Exception:
        df = pd.read_excel(input_file, sheet_name=0)
        print(f"   [OK] Membaca sheet pertama (fallback)")

df_original = df.copy()
print(f"   ✓ Total data awal: {len(df)} bahan makanan\n")

# ============================================================
# TAHAP PEMBERSIHAN (dari pembersihan.py)
# ============================================================
print("[PEMBERSIHAN] Memulai proses pembersihan data...\n")

# -------- STEP 1: Hapus kategori 'lainnya' --------
print("[STEP 1] Menghapus kategori 'lainnya'...")
if 'kategori' in df.columns:
    df_cleaned = df[df['kategori'] != 'lainnya']
    removed_lainnya = len(df) - len(df_cleaned)
    print(f"   - Baris yang dihapus: {removed_lainnya}")
    print(f"   - Total setelah: {len(df_cleaned)}\n")
else:
    df_cleaned = df.copy()
    removed_lainnya = 0
    print("   ⚠️ Kolom 'kategori' tidak ditemukan, skip step ini\n")

# -------- STEP 2: Hapus bahan berdasarkan kata kunci --------
print("[STEP 2] Menghapus bahan makanan berdasarkan kata kunci...")
before_keyword = len(df_cleaned)
for kata in KEYWORDS_TO_REMOVE:
    if 'name' in df_cleaned.columns:
        df_cleaned = df_cleaned[~df_cleaned['name'].str.contains(kata, case=False, na=False)]

removed_keywords = before_keyword - len(df_cleaned)
print(f"   - Baris yang dihapus: {removed_keywords}")
print(f"   - Total setelah: {len(df_cleaned)}\n")

# -------- STEP 3: Strip kata tertentu dari nama bahan --------
print("[STEP 3] Menghilangkan kata dari nama bahan (strip words)...")
print(f"   Kata yang dihilangkan: {STRIP_WORDS}")
stripped_count = 0

if 'name' in df_cleaned.columns:
    new_names = []
    for idx, row in df_cleaned.iterrows():
        original_name = row['name']
        cleaned_name, removed_words = strip_words_from_name(original_name)
        if removed_words:
            print(f"     [STRIP] '{original_name}' -> '{cleaned_name}' (dihapus: {', '.join(removed_words)})")
            stripped_count += 1
        new_names.append(cleaned_name)
    
    df_cleaned = df_cleaned.copy()
    df_cleaned['name'] = new_names
    print(f"   - Nama yang diubah: {stripped_count}")
    print(f"   - Total data: {len(df_cleaned)}\n")
else:
    print("   ⚠️ Kolom 'name' tidak ditemukan, skip step ini\n")

# -------- STEP 4: Hapus duplikat --------
print("[STEP 4] Menghapus data duplikat (nama yang sama)...")
if 'name' in df_cleaned.columns:
    df_cleaned['_name_lower'] = df_cleaned['name'].str.lower().str.strip()
    duplicated_mask = df_cleaned.duplicated(subset='_name_lower', keep='first')
    duplicated_names = df_cleaned.loc[duplicated_mask, 'name'].tolist()
    
    if duplicated_names:
        print(f"   - Duplikat ditemukan ({len(duplicated_names)}):")
        for dup_name in duplicated_names:
            print(f"     x {dup_name}")
    else:
        print("   - Tidak ada duplikat ditemukan")
    
    before_dedup = len(df_cleaned)
    df_cleaned = df_cleaned.drop_duplicates(subset='_name_lower', keep='first')
    removed_duplicates = before_dedup - len(df_cleaned)
    df_cleaned = df_cleaned.drop(columns=['_name_lower'])
    
    print(f"   - Baris duplikat dihapus: {removed_duplicates}")
    print(f"   - Total setelah: {len(df_cleaned)}\n")
else:
    removed_duplicates = 0
    print("   ⚠️ Kolom 'name' tidak ditemukan, skip step ini\n")

# ============================================================
# TAHAP PEMBAGIAN (dari pembagian.py)
# ============================================================
print("[PEMBAGIAN] Membagi data per kategori...\n")

# Cek kolom kategori
if 'kategori' not in df_cleaned.columns:
    print("   ❌ Kolom 'kategori' tidak ditemukan!")
    exit(1)

# Tampilkan semua kategori yang tersedia
all_categories = df_cleaned['kategori'].unique()
print(f"   [INFO] Kategori yang tersedia: {sorted(all_categories)}")
print(f"   [INFO] Kategori yang akan diambil: {TARGET_CATEGORIES}\n")

# -------- Filter dan simpan per kategori --------
print("[STEP] Menulis ke sheet lauk & sayuran...")

with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    for kategori in TARGET_CATEGORIES:
        df_kategori = df_cleaned[df_cleaned['kategori'].str.lower() == kategori.lower()]
        
        if len(df_kategori) == 0:
            print(f"   ⚠️ Kategori '{kategori}' tidak ditemukan dalam data, skip")
            continue
        
        sheet_name = kategori
        df_kategori.to_excel(writer, sheet_name=sheet_name, index=False)
        format_excel_sheet(writer.sheets[sheet_name])
        
        print(f"   ✅ Sheet '{sheet_name}' dibuat: {len(df_kategori)} bahan makanan")

# -------- Hapus sheet 'Data Cleaned' jika ada --------
print("[CLEANUP] Menghapus sheet 'Data Cleaned' (intermediate)...")
try:
    from openpyxl import load_workbook
    wb = load_workbook(input_file)
    if 'Data Cleaned' in wb.sheetnames:
        del wb['Data Cleaned']
        wb.save(input_file)
        print("   ✓ Sheet 'Data Cleaned' berhasil dihapus")
    else:
        print("   ℹ️ Sheet 'Data Cleaned' tidak ditemukan, skip")
except Exception as e:
    print(f"   ⚠️ Gagal menghapus sheet 'Data Cleaned': {e}")
finally:
    wb.close()

# ============================================================
# RINGKASAN
# ============================================================
print("\n" + "=" * 80)
print("[RESULT] RINGKASAN PEMBERSIHAN + PEMBAGIAN DATA")
print("=" * 80)
print(f"\nData Awal       : {len(df_original)} bahan makanan")
print(f"Hapus 'lainnya' : -{removed_lainnya} baris")
print(f"Hapus keyword   : -{removed_keywords} baris")
print(f"Strip words     : {stripped_count} nama diubah")
print(f"Hapus duplikat  : -{removed_duplicates} baris")
print(f"Data Setelah Pembersihan: {len(df_cleaned)} bahan makanan")
print(f"Total dihapus   : {len(df_original) - len(df_cleaned)} baris ({((len(df_original) - len(df_cleaned))/len(df_original)*100):.1f}%)")

if 'kategori' in df_cleaned.columns:
    print(f"\n[KATEGORI] Kategori yang tersisa:")
    kategori_counts = df_cleaned['kategori'].value_counts()
    for cat, count in kategori_counts.items():
        print(f"   - {cat}: {count} items")

print(f"\n[OUTPUT] Hasil pembagian:")
total_exported = 0
for kategori in TARGET_CATEGORIES:
    df_kat = df_cleaned[df_cleaned['kategori'].str.lower() == kategori.lower()]
    count = len(df_kat)
    total_exported += count
    print(f"   -> Sheet '{kategori}': {count} bahan makanan")

print(f"\nTotal data diekspor: {total_exported} bahan makanan")
print(f"Kategori lain (tidak diekspor): {len(df_cleaned) - total_exported} bahan makanan")

print("\n[DONE] STEP 3-4 SELESAI: Pembersihan + Pembagian berhasil disimpan!")
print(f"📁 File: {input_file}")
print(f"📊 Output sheets: {', '.join(TARGET_CATEGORIES)}")
print("\n" + "=" * 80)
print(f"📁 File: {input_file}")
print("\n" + "=" * 80)
print("PIPELINE SELESAI!")
print("Sheet yang tersedia di nutrition_pipeline.xlsx:")
print("   1. 'Dataset Asli'        - Data original dari CSV")
print("   2. 'Nutrition Scaled'    - Data setelah Min-Max Scaling")
print("   3. 'One-Hot Encoded'     - Data setelah One-Hot Encoding")
print("   4. 'lauk'                - Data kategori lauk")
print("   5. 'sayuran'             - Data kategori sayuran")
print("=" * 80)